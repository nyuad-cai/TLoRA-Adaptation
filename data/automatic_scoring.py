"""
auto-label LLM predictions that exactly match the ground-truth reference.
"""

import sys
import re
import unicodedata
from pathlib import Path

import pandas as pd

MERGED_COLS = ["ID", "QUESTION", "GROUND TRUTH"]
MODEL_COL = "MODEL"
PREDICTION_COL = "PREDICTION"
LABEL_COL = "LABEL"
CORRECT_LABEL = "correct"

NORMALIZE_TA_MARBUTA = False
_TASHKEEL = re.compile(r"[\u0617-\u061A\u064B-\u0652\u0670\u06D6-\u06ED]")
_TATWEEL = "\u0640"
_WHITESPACE = re.compile(r"\s+")


def normalize_arabic(text: str) -> str:
    """Normalize Arabic text for exact-match comparison"""
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)

    text = unicodedata.normalize("NFC", text)
    text = _TASHKEEL.sub("", text)
    text = text.replace(_TATWEEL, "")

    for ch in ("\u0623", "\u0625", "\u0622", "\u0671"):  
        text = text.replace(ch, "\u0627")  

    text = text.replace("\u0649", "\u064a")  

    if NORMALIZE_TA_MARBUTA:
        text = text.replace("\u0629", "\u0647")  

    text = _WHITESPACE.sub(" ", text).strip()

    return text


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    in_path = Path(sys.argv[1])
    if not in_path.is_file():
        print(f"File not found: {in_path}")
        sys.exit(1)

    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else in_path.with_name(
        in_path.stem + "_labeled.csv"
    )

    df = pd.read_csv(in_path, dtype=str, keep_default_na=False)

    missing = [c for c in MERGED_COLS + [MODEL_COL, PREDICTION_COL] if c not in df.columns]
    if missing:
        print(f"Missing required columns: {missing}")
        print(f"Found columns: {list(df.columns)}")
        sys.exit(1)

    if LABEL_COL not in df.columns:
        df[LABEL_COL] = ""

    for col in MERGED_COLS:
        df[col] = df[col].replace("", pd.NA).ffill().fillna("")

    gt_norm = df["GROUND TRUTH"].map(normalize_arabic)
    pred_norm = df[PREDICTION_COL].map(normalize_arabic)

    both_nonempty = (gt_norm != "") & (pred_norm != "")
    matches = both_nonempty & (gt_norm == pred_norm)

    empty_label = df[LABEL_COL].fillna("").str.strip() == ""
    to_fill = matches & empty_label

    df.loc[to_fill, LABEL_COL] = CORRECT_LABEL

    total_rows = len(df)
    total_matches = int(matches.sum())
    newly_labeled = int(to_fill.sum())
    already_labeled_matches = int((matches & ~empty_label).sum())
    n_questions = df["ID"].nunique()

    print(f"\n=== Summary ===")
    print(f"Total rows:             {total_rows}")
    print(f"Unique questions:       {n_questions}")
    print(f"Exact matches found:    {total_matches}")
    print(f"  -> newly labeled:     {newly_labeled}")
    print(f"  -> already had label: {already_labeled_matches}")
    print(f"Rows left for humans:   {total_rows - int((df[LABEL_COL].str.strip() != '').sum())}")

    print(f"\n=== Exact matches by model ===")
    per_model = (
        df.assign(_match=matches)
          .groupby(MODEL_COL)["_match"]
          .agg(["sum", "count"])
          .rename(columns={"sum": "matches", "count": "total"})
          .sort_values("matches", ascending=False)
    )
    per_model["pct"] = (per_model["matches"] / per_model["total"] * 100).round(1)
    print(per_model.to_string())

    first_in_group = df["ID"] != df["ID"].shift()
    for col in MERGED_COLS:
        df.loc[~first_in_group, col] = ""

    if out_path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "annotations"

        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=1, column=j, value=col)

        full = df.copy()
        for col in MERGED_COLS:
            full[col] = full[col].replace("", pd.NA).ffill().fillna("")

        for i, row in enumerate(full.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        group_starts = full.index[first_in_group.values].tolist() + [len(full)]
        col_letters = {c: get_column_letter(list(df.columns).index(c) + 1) for c in MERGED_COLS}
        for a, b in zip(group_starts[:-1], group_starts[1:]):
            if b - a > 1:
                for c, letter in col_letters.items():
                    ws.merge_cells(f"{letter}{a+2}:{letter}{b+1}")

        wb.save(out_path)
    else:
        df.to_csv(out_path, index=False, encoding="utf-8-sig")
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()